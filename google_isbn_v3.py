import json
import requests
from pprint import pprint  # подключили Pprint для красоты выдачи текста
from openpyxl import Workbook
import openpyxl
from datetime import datetime
url = "https://www.googleapis.com/books/v1/volumes?q=isbn:"
url2 = 'https://www.googleapis.com/books/v1/volumes/'

def add_to_exel(dict_to_exel):
    tab_headers_dict = {
        'Название': 'title',
        'Автор': 'authors',
        'Издательство': 'publisher',
        'Город': '',
        'Год': 'publishedDate',
        'Страницы': 'pageCount',
        'Переплет': '',
        'Цена': '',
        'Закуп': '',
        'Количество': '',
        'Жанр': 'categories',
        'Описание': 'description',
        'ISBN': 'industryIdentifiers'
    }
    for header in list(tab_headers_dict):
        try:
            # print(tab_headers_dict[header])
            tab_headers_dict[header] = dict_to_exel[tab_headers_dict[header]]
            # print(tab_headers_dict[header])
        except Exception as e:
            pass
            tab_headers_dict[header] = ''
            # print(header)
    try:
        excel_file = openpyxl.load_workbook('new_baza.xlsx')  # пытаемся открыть файл
        excel_sheet = excel_file['list1']
    except FileNotFoundError:  # а если его нет, создаем новый
        print('Создан новый файл')
        excel_file = Workbook()
        excel_sheet = excel_file.create_sheet(title='list1', index=0)
        excel_sheet['A1'] = 'Название'
        excel_sheet['B1'] = 'Автор'
        excel_sheet['C1'] = 'Издательство'
        excel_sheet['D1'] = 'Город'
        excel_sheet['E1'] = 'Год'
        excel_sheet['F1'] = 'Страницы'
        excel_sheet['G1'] = 'Переплет'
        excel_sheet['H1'] = 'Цена'
        excel_sheet['I1'] = 'Закуп'
        excel_sheet['J1'] = 'Количество'
        excel_sheet['K1'] = 'Жанр'
        excel_sheet['L1'] = 'Описание'
    a = tab_headers_dict
    book_row = (a['Название'], ' '.join(a['Автор']), a['Издательство'], a['Город'], a['Год'], a['Страницы'], a['Переплет'], a['Цена'], a['Закуп'], a['Количество'], ' '.join(a['Жанр']), a['Описание'], )
    excel_sheet.append(book_row)
    excel_file.save('new_baza.xlsx')
    print(book_row)

while True:
    isbn = input("Введите ISBN: ").strip()
    #isbn = '9785855971194'
    resp = requests.get(url + isbn)
    book_data = json.loads(resp.text)

    totalItems = int(book_data['totalItems'])
    if totalItems == 0:
        print('Ничего не найдено')
        continue

    id = book_data['items'][0]['id']
    selfLink = book_data['items'][0]['selfLink']
    book_list_1 = list(book_data['items'][0]['volumeInfo'])         #list() получаем список ключей словаря

    book_list_dict_1 = {'id': book_data['items'][0]['id']}          #создаем свой словарь
    for atribut in book_list_1:
        book_list_dict_1[atribut] = book_data['items'][0]['volumeInfo'][atribut]    #заполняем свой словарь, ключи и значения копируются из исходника
    resp2 = requests.get(selfLink)
    book_data2 = json.loads(resp2.text)
    book_list_2 = list(book_data2['volumeInfo'])
    book_list_dict_2 = {'id': book_data2['id']}
    for atribut2 in book_list_2:
        book_list_dict_2[atribut2] = book_data2['volumeInfo'][atribut2]
    try:
        with open(str(isbn) + '_-_' + datetime.today().strftime("%m_%d_%H_%M_%S") + '.json', 'x', encoding='UTF-8') as log_file:
            json.dump(book_list_dict_2, log_file)
            print('Данные сохранены в файл')
    except FileExistsError:
            print('Не удалось записать данные в файл')
    add_to_exel(book_list_dict_2)
